# FICHIERS EXCEL
# .XLSX
# openpyxl
import openpyxl

wb1 = openpyxl.load_workbook("octobre.xlsx", data_only=True)
wb2 = openpyxl.load_workbook("novembre.xlsx", data_only=True)
wb3 = openpyxl.load_workbook("decembre.xlsx", data_only=True)

# {"Pommes" : (760, 660, 900)}

def ajouter_data_depuis_wb(wb, d):
    sheet = wb.active
    for row in range(2, sheet.max_row):
        nom_article = sheet.cell(row, 1).value
        if not nom_article:
            break
        total_ventes = sheet.cell(row, 4).value
        if d.get(nom_article):
            d[nom_article].append(total_ventes)
        else:
            d[nom_article] = [total_ventes]

donnees = {}
ajouter_data_depuis_wb(wb1, donnees)
ajouter_data_depuis_wb(wb2, donnees)
ajouter_data_depuis_wb(wb3, donnees)


print(donnees)
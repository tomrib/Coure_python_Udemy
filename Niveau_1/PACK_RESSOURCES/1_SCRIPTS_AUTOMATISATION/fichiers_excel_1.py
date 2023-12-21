# FICHIERS EXCEL
# .XLSX
# openpyxl
import openpyxl

wb = openpyxl.load_workbook("octobre.xlsx")
print(wb.sheetnames)
#sheet = wb[wb.sheetnames[0]]
sheet = wb.active

cell = sheet["B7"]
print(cell.value)

print(sheet.max_row)
print(sheet.max_column)
#for row in range(2, 7):
#    cell = sheet.cell(row, 2)
#    print(cell.value)